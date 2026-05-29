#!/usr/bin/env python3
"""Genererer en Excel-fil med all brukervendt tekst fra straverso.com.

Kolonner: ID | Seksjon | Element | Dagens tekst | Ny tekst
ID-en mapper tilbake til kildekoden, slik at endringer kan plasseres presist.
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from content_data import ROWS as rows

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

wb = Workbook()
ws = wb.active
ws.title = "Innhold"

headers = ["ID", "Seksjon", "Element", "Dagens tekst", "Ny tekst"]

# Styling
header_fill = PatternFill("solid", fgColor="0F1354")  # indigo
header_font = Font(bold=True, color="F5F2EB", size=11)  # offwhite
new_col_fill = PatternFill("solid", fgColor="FFF4F4")   # svak coral
id_font = Font(color="888888", size=9)
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
top_wrap = Alignment(vertical="top", wrap_text=True)

# Header-rad
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(vertical="center", horizontal="left")
    c.border = border

# Datarader
for i, (rid, section, element, current) in enumerate(rows, start=2):
    ws.cell(row=i, column=1, value=rid).font = id_font
    ws.cell(row=i, column=2, value=section)
    ws.cell(row=i, column=3, value=element)
    ws.cell(row=i, column=4, value=current)
    ws.cell(row=i, column=5, value="")  # Ny tekst – fylles av bruker
    for col in range(1, 6):
        cell = ws.cell(row=i, column=col)
        cell.alignment = top_wrap
        cell.border = border
        if col == 5:
            cell.fill = new_col_fill

# Kolonnebredder
widths = [22, 14, 30, 70, 70]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Frys header + de tre første kolonnene for enkel navigering
ws.freeze_panes = "D2"
ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

out = os.path.join(ROOT, "straverso-innhold.xlsx")
wb.save(out)
print(f"Skrev {out} med {len(rows)} tekstrader.")
